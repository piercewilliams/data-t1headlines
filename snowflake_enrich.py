#!/usr/bin/env python3
"""
snowflake_enrich.py — Build data/snowflake_enrichment.json from TRACKER_ENRICHED.

Reads all of TRACKER_ENRICHED (~2K rows, built by ops-hub/scripts/model_tracker.py)
via the headless service account and produces two output sections:

  articles  — per-article O&O traffic, IAB topic, similarity, median comparison
              keyed by normalized published URL
  authors   — per-author aggregates broken down by publication and IAB topic,
              computed from all articles in TRACKER_ENRICHED for that author

generate_site.py loads this file to power per-author pub breakdowns, topic-stratified
formula analysis, and richer editorial playbooks.

Authentication: RSA key-pair (headless — same auth as model_tracker.py)

Env vars (optional):
    SNOWFLAKE_PRIVATE_KEY_PATH   path to RSA private key (.p8 file)

Usage:
    python3 snowflake_enrich.py [--out FILE]


Also audits Tarrow's XLSX for articles that appear on syndication platforms but have
no matching row in Sara's tracker. These are logged to data/tracker_gaps.json —
articles the team produced and syndicated but never formally tracked.

Usage:
    python3 snowflake_enrich.py [--out FILE] [--tarrow FILE]
"""

import argparse
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

SF_ACCOUNT  = "wvb49304-mcclatchy_eval"
SF_USER     = "GROWTH_AND_STRATEGY_SERVICE_USER"
SF_KEY_PATH = os.getenv(
    "SNOWFLAKE_PRIVATE_KEY_PATH",
    str(Path.home() / ".credentials" / "growth_strategy_service_rsa_key.p8"),
)

DEFAULT_OUT    = "data/snowflake_enrichment.json"
DEFAULT_TARROW = "Top Stories 2026 Syndication.xlsx"
DEFAULT_GAPS   = "data/tracker_gaps.json"

FETCH_SQL = """
SELECT
    LOWER(RTRIM(
        CASE WHEN PUBLISHED_URL LIKE '%://amp.%'
             THEN REPLACE(PUBLISHED_URL, '://amp.', '://www.')
             ELSE PUBLISHED_URL END,
        '/'))                       AS url_norm,
    PUBLISHED_URL,
    AUTHOR,
    HEADLINE,
    article_domain                  AS domain,
    word_count,
    primary_iab_topic               AS iab_topic,
    total_pvs,
    search_pvs,
    social_pvs,
    direct_pvs,
    applenews_pvs,
    smartnews_pvs,
    newsbreak_pvs,
    subscriber_pvs,
    article_vs_co_median,
    is_hit,
    cluster_id,
    cluster_avg_sim_desc,
    cluster_total_pvs,
    cluster_vs_co_median,
    cluster_hit_rate
FROM MCC_PRESENTATION.CONTENT_SCALING_AGENT.TRACKER_ENRICHED
WHERE PUBLISHED_URL IS NOT NULL
QUALIFY ROW_NUMBER() OVER (PARTITION BY url_norm ORDER BY total_pvs DESC NULLS LAST) = 1
"""

COLS = [
    "url_norm", "published_url", "author", "headline", "domain", "word_count",
    "iab_topic", "total_pvs", "search_pvs", "social_pvs", "direct_pvs",
    "applenews_pvs", "smartnews_pvs", "newsbreak_pvs", "subscriber_pvs",
    "article_vs_co_median", "is_hit",
    "cluster_id", "cluster_avg_sim_desc", "cluster_total_pvs",
    "cluster_vs_co_median", "cluster_hit_rate",
]


def get_conn():
    from cryptography.hazmat.primitives.serialization import (
        load_pem_private_key, Encoding, PrivateFormat, NoEncryption,
    )
    import snowflake.connector

    raw = Path(SF_KEY_PATH).read_bytes()
    key = load_pem_private_key(raw, password=None)
    der = key.private_bytes(Encoding.DER, PrivateFormat.PKCS8, NoEncryption())
    return snowflake.connector.connect(
        account=SF_ACCOUNT,
        user=SF_USER,
        private_key=der,
        warehouse="GROWTH_AND_STRATEGY_ROLE_WH",
        role="GROWTH_AND_STRATEGY_ROLE",
    )


def _f(v):
    """Coerce Decimal/None → float or None."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _i(v):
    """Coerce to int or None."""
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def fetch_rows(cur):
    cur.execute(FETCH_SQL)
    rows = cur.fetchall()
    return [dict(zip(COLS, r)) for r in rows]


def build_articles(rows):
    """Per-article enrichment dict keyed by normalized URL."""
    articles = {}
    for r in rows:
        key = r["url_norm"]
        if not key:
            continue
        articles[key] = {
            "author":               r["author"] or "",
            "headline":             r["headline"] or "",
            "domain":               r["domain"] or "",
            "word_count":           _i(r["word_count"]),
            "iab_topic":            r["iab_topic"] or None,
            "total_pvs":            _i(r["total_pvs"]),
            "search_pvs":           _i(r["search_pvs"]),
            "social_pvs":           _i(r["social_pvs"]),
            "direct_pvs":           _i(r["direct_pvs"]),
            "applenews_pvs":        _i(r["applenews_pvs"]),
            "smartnews_pvs":        _i(r["smartnews_pvs"]),
            "newsbreak_pvs":        _i(r["newsbreak_pvs"]),
            "subscriber_pvs":       _i(r["subscriber_pvs"]),
            "article_vs_co_median": _f(r["article_vs_co_median"]),
            "is_hit":               _i(r["is_hit"]),
            "cluster_id":           r["cluster_id"] or None,
            "cluster_avg_sim_desc": _f(r["cluster_avg_sim_desc"]),
            "cluster_total_pvs":    _i(r["cluster_total_pvs"]),
            "cluster_vs_co_median": _f(r["cluster_vs_co_median"]),
            "cluster_hit_rate":     _f(r["cluster_hit_rate"]),
        }
    return articles


def build_authors(rows):
    """
    Per-author aggregates computed from all TRACKER_ENRICHED rows for that author.

    Structure:
        author_name → {
            total_articles, total_pvs, avg_pvs, hit_rate,
            by_pub:   { domain → {articles, total_pvs, avg_pvs, hits, hit_rate} }
            by_topic: { iab_topic → {articles, total_pvs, avg_pvs, hits, hit_rate} }
        }
    """
    # Accumulate raw counts per author → (pub|topic) → bucket
    # bucket = [total_pvs, article_count, hit_count]
    Author = lambda: {
        "pvs":     0,
        "n":       0,
        "hits":    0,
        "by_pub":   defaultdict(lambda: [0, 0, 0]),
        "by_topic": defaultdict(lambda: [0, 0, 0]),
    }
    acc = defaultdict(Author)

    for r in rows:
        author = (r["author"] or "").strip()
        if not author:
            continue
        pvs   = _i(r["total_pvs"]) or 0
        hit   = _i(r["is_hit"])
        pub   = r["domain"] or "unknown"
        topic = r["iab_topic"] or "Unknown"

        a = acc[author]
        a["pvs"]  += pvs
        a["n"]    += 1
        if hit is not None:
            a["hits"] += hit

        a["by_pub"][pub][0]   += pvs
        a["by_pub"][pub][1]   += 1
        if hit is not None:
            a["by_pub"][pub][2] += hit

        a["by_topic"][topic][0]   += pvs
        a["by_topic"][topic][1]   += 1
        if hit is not None:
            a["by_topic"][topic][2] += hit

    def _bucket(raw_dict):
        out = {}
        for key, (pvs, n, hits) in raw_dict.items():
            out[key] = {
                "articles":  n,
                "total_pvs": pvs,
                "avg_pvs":   round(pvs / n) if n else None,
                "hit_rate":  round(hits / n, 3) if n else None,
            }
        return out

    authors = {}
    for author, a in acc.items():
        n = a["n"]
        authors[author] = {
            "total_articles": n,
            "total_pvs":      a["pvs"],
            "avg_pvs":        round(a["pvs"] / n) if n else None,
            "hit_rate":       round(a["hits"] / n, 3) if n else None,
            "by_pub":         _bucket(a["by_pub"]),
            "by_topic":       _bucket(a["by_topic"]),
        }

    return authors


def _norm(url):
    """Normalize URL for matching: lowercase, strip trailing slash, fix amp."""
    if not url:
        return ""
    url = str(url).strip().lower()
    url = re.sub(r'://amp\.', '://www.', url)
    return url.rstrip('/')


def collect_tarrow_urls(tarrow_file):
    """
    Read all O&O article URLs out of Tarrow's syndication XLSX.

    Returns a list of dicts:
        { url_norm, raw_url, headline, author, platform, date, platform_views }

    Covers Apple News (Publisher Article ID), SmartNews (url).
    MSN and Yahoo have no reliable O&O URL field so are skipped.
    """
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed — skipping Tarrow gap audit.")
        return []

    if not Path(tarrow_file).exists():
        print(f"  {tarrow_file} not found — skipping Tarrow gap audit.")
        return []

    records = []
    wb = openpyxl.load_workbook(tarrow_file, read_only=True, data_only=True)

    # ── Apple News ────────────────────────────────────────────────────────────
    if "Apple News" in wb.sheetnames:
        ws = wb["Apple News"]
        hdrs = [str(c) if c else "" for c in
                next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(hdrs)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_url = row[idx["Publisher Article ID"]] if "Publisher Article ID" in idx else None
            if not raw_url or not str(raw_url).startswith("http"):
                continue
            records.append({
                "url_norm":       _norm(raw_url),
                "raw_url":        str(raw_url).strip(),
                "headline":       str(row[idx["Article"]]).strip() if "Article" in idx else "",
                "author":         str(row[idx["Author"]]).strip() if "Author" in idx else "",
                "platform":       "Apple News",
                "date":           str(row[idx["Date Published"]])[:10] if "Date Published" in idx else "",
                "platform_views": row[idx["Total Views"]] if "Total Views" in idx else None,
            })

    # ── SmartNews ─────────────────────────────────────────────────────────────
    if "SmartNews" in wb.sheetnames:
        ws = wb["SmartNews"]
        hdrs = [str(c) if c else "" for c in
                next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(hdrs)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_url = row[idx["url"]] if "url" in idx else None
            if not raw_url or not str(raw_url).startswith("http"):
                continue
            records.append({
                "url_norm":       _norm(raw_url),
                "raw_url":        str(raw_url).strip(),
                "headline":       str(row[idx["title"]]).strip() if "title" in idx else "",
                "author":         "",
                "platform":       "SmartNews",
                "date":           str(row[idx["month"]])[:10] if "month" in idx else "",
                "platform_views": row[idx["article_view"]] if "article_view" in idx else None,
            })

    print(f"  Read {len(records)} Tarrow article records "
          f"(AN={sum(1 for r in records if r['platform']=='Apple News')}, "
          f"SN={sum(1 for r in records if r['platform']=='SmartNews')}).")
    return records


def build_gaps(tarrow_records, tracker_urls):
    """
    Find Tarrow URLs that have no matching row in Sara's tracker.

    tracker_urls  — set of normalized URLs from TRACKER_ENRICHED
    Returns list of gap dicts sorted by platform_views descending.
    """
    # Deduplicate Tarrow records by url_norm, keeping the highest-view entry
    best = {}
    for r in tarrow_records:
        key = r["url_norm"]
        if not key:
            continue
        if key not in best or (r["platform_views"] or 0) > (best[key]["platform_views"] or 0):
            best[key] = r

    gaps = []
    for key, r in best.items():
        if key not in tracker_urls:
            gaps.append({
                "url":            r["raw_url"],
                "headline":       r["headline"],
                "author":         r["author"],
                "platform":       r["platform"],
                "date":           r["date"],
                "platform_views": int(r["platform_views"]) if r["platform_views"] else None,
            })

    gaps.sort(key=lambda g: g["platform_views"] or 0, reverse=True)
    return gaps


def main():
    parser = argparse.ArgumentParser(description="Build snowflake_enrichment.json")
    parser.add_argument("--out",    default=DEFAULT_OUT,
                        help=f"Output JSON path (default: {DEFAULT_OUT})")
    parser.add_argument("--tarrow", default=DEFAULT_TARROW,
                        help=f"Tarrow XLSX path (default: {DEFAULT_TARROW})")
    parser.add_argument("--gaps",   default=DEFAULT_GAPS,
                        help=f"Gap report output path (default: {DEFAULT_GAPS})")
    args = parser.parse_args()

    print("Connecting to Snowflake…")
    conn = get_conn()
    cur  = conn.cursor()
    print("Connected.")

    print("Fetching TRACKER_ENRICHED…")
    rows = fetch_rows(cur)
    print(f"  {len(rows)} rows fetched.")

    cur.close()
    conn.close()

    print("Building article index…")
    articles = build_articles(rows)
    print(f"  {len(articles)} articles indexed.")

    print("Building author aggregates…")
    authors = build_authors(rows)
    print(f"  {len(authors)} authors indexed.")

    out = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "row_count": len(rows),
        "articles":  articles,
        "authors":   authors,
    }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out, indent=2, default=str))
    size_kb = out_path.stat().st_size // 1024
    print(f"\nWrote {args.out} ({size_kb} KB, {len(articles)} articles, {len(authors)} authors).")

    # ── Tracker gap audit ─────────────────────────────────────────────────────
    print(f"\nAuditing Tarrow → tracker gaps ({args.tarrow})…")
    tarrow_records = collect_tarrow_urls(args.tarrow)
    if tarrow_records:
        tracker_urls = set(articles.keys())
        gaps = build_gaps(tarrow_records, tracker_urls)

        gaps_out = {
            "generated":     datetime.now(timezone.utc).isoformat(),
            "tarrow_total":  len({r["url_norm"] for r in tarrow_records if r["url_norm"]}),
            "tracker_total": len(tracker_urls),
            "gap_count":     len(gaps),
            "gaps":          gaps,
        }
        gaps_path = Path(args.gaps)
        gaps_path.parent.mkdir(parents=True, exist_ok=True)
        gaps_path.write_text(json.dumps(gaps_out, indent=2, default=str))

        print(f"  {len(gaps)} untracked articles found out of "
              f"{gaps_out['tarrow_total']} unique Tarrow URLs.")
        if gaps:
            print(f"  Top 5 by platform views:")
            for g in gaps[:5]:
                views = f"{g['platform_views']:,}" if g['platform_views'] else "?"
                print(f"    [{g['platform']}] {views} views — {g['headline'][:60]}")
        print(f"  Full gap report → {args.gaps}")


if __name__ == "__main__":
    main()
