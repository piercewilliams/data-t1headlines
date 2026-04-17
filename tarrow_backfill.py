#!/usr/bin/env python3
"""
tarrow_backfill.py — Backfill Sara's tracker Syndication platform column from Tarrow data.

Pipeline position: AFTER download_tarrow.py, BEFORE snowflake_enrich.py + generate_site.py.

Logic:
  For each tracker row with an empty Syndication platform:
    - If Published URL/Link matches Apple News Publisher Article ID → "Apple News"
    - If it matches SmartNews url → "Smart News"
    - If both → "Apple News, Smart News"

Never overwrites a cell that already contains a value.
Exports Sara's tracker as Tracker Template.xlsx for generate_site.py.
Writes data/tarrow_backfill_report.json.

Environment:
    GOOGLE_SERVICE_ACCOUNT_FILE  — path to service account JSON file
    GOOGLE_SERVICE_ACCOUNT_JSON  — raw JSON string (fallback)
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

TARROW_XLSX = "Top Stories 2026 Syndication.xlsx"
TRACKER_SHEET_ID = "14_0eK46g3IEj7L_yp9FIdWwvnuYI5f-vAuP7DDhSPg8"
TRACKER_XLSX_OUT = "Tracker Template.xlsx"
REPORT_PATH = Path("data/tarrow_backfill_report.json")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

DRIVE_EXPORT_URL = (
    "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
)


def _load_credentials():
    try:
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("Error: google-auth is not installed.")
        sys.exit(1)

    sa_file = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    if sa_file and Path(sa_file).exists():
        return Credentials.from_service_account_file(sa_file, scopes=SCOPES)

    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if sa_json:
        try:
            info = json.loads(sa_json)
        except json.JSONDecodeError as exc:
            print(f"Error: GOOGLE_SERVICE_ACCOUNT_JSON is not valid JSON: {exc}")
            sys.exit(1)
        return Credentials.from_service_account_info(info, scopes=SCOPES)

    # Local dev fallback
    local = Path.home() / ".credentials" / "pierce-tools.json"
    if local.exists():
        return Credentials.from_service_account_file(str(local), scopes=SCOPES)

    print(
        "Error: no service account credentials found.\n"
        "Set GOOGLE_SERVICE_ACCOUNT_FILE or GOOGLE_SERVICE_ACCOUNT_JSON."
    )
    sys.exit(1)


def _norm(url: str) -> str:
    """Normalize a URL for comparison: lowercase, strip trailing slash."""
    if not url:
        return ""
    return url.strip().lower().rstrip("/")


def build_tarrow_platform_map(xlsx_path: str) -> dict:
    """
    Read Tarrow XLSX and return {norm_url: set_of_platforms}.
    Apple News Publisher Article ID → "Apple News"
    SmartNews url → "Smart News"
    """
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is not installed.")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    platform_map: dict = defaultdict(set)

    # ── Apple News ────────────────────────────────────────────────────────────
    if "Apple News" in wb.sheetnames:
        ws_an = wb["Apple News"]
        rows = ws_an.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        try:
            pai_idx = headers.index("Publisher Article ID")
        except ValueError:
            print("Warning: 'Publisher Article ID' column not found in Apple News tab.")
            pai_idx = None

        if pai_idx is not None:
            an_count = 0
            for row in rows:
                if len(row) <= pai_idx:
                    continue
                url = row[pai_idx]
                if not url:
                    continue
                norm = _norm(str(url))
                if norm:
                    platform_map[norm].add("Apple News")
                    an_count += 1
            print(f"  Apple News: {an_count} URLs loaded")

    # ── SmartNews ─────────────────────────────────────────────────────────────
    if "SmartNews" in wb.sheetnames:
        ws_sn = wb["SmartNews"]
        rows = ws_sn.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        try:
            url_idx = headers.index("url")
        except ValueError:
            print("Warning: 'url' column not found in SmartNews tab.")
            url_idx = None

        if url_idx is not None:
            sn_count = 0
            for row in rows:
                if len(row) <= url_idx:
                    continue
                url = row[url_idx]
                if not url or not str(url).startswith("http"):
                    continue
                norm = _norm(str(url))
                if norm:
                    platform_map[norm].add("Smart News")
                    sn_count += 1
            print(f"  SmartNews: {sn_count} URLs loaded")

    wb.close()
    print(f"  Total unique Tarrow URLs: {len(platform_map)}")
    return dict(platform_map)


def _platform_label(platforms: set) -> str:
    """Convert a set of platform names to a single cell value."""
    if platforms == {"Apple News"}:
        return "Apple News"
    if platforms == {"Smart News"}:
        return "Smart News"
    if "Apple News" in platforms and "Smart News" in platforms:
        return "Apple News, Smart News"
    # Fallback for any unexpected combination
    return ", ".join(sorted(platforms))


def compute_fills(all_values: list, url_idx: int, plat_idx: int,
                  tarrow_map: dict) -> list:
    """
    Return list of (sheet_row_1indexed, platform_label) for cells to fill.
    Only rows where Syndication platform is empty and URL matches Tarrow.
    """
    fills = []
    for i, row in enumerate(all_values[1:], start=2):  # row 1 = header
        url = row[url_idx].strip() if len(row) > url_idx else ""
        plat = row[plat_idx].strip() if len(row) > plat_idx else ""

        if not url or plat:  # skip if no URL or platform already set
            continue

        norm = _norm(url)
        if norm in tarrow_map:
            label = _platform_label(tarrow_map[norm])
            fills.append((i, label))

    return fills


def apply_fills(ws, fills: list, plat_col: int) -> None:
    """Batch-write platform values to Google Sheet."""
    if not fills:
        print("  No cells to fill.")
        return

    import gspread

    updates = []
    for row_idx, label in fills:
        a1 = gspread.utils.rowcol_to_a1(row_idx, plat_col)
        updates.append({"range": a1, "values": [[label]]})

    # gspread batch_update limit is 1000 ranges; chunk if needed
    chunk_size = 500
    for i in range(0, len(updates), chunk_size):
        ws.spreadsheet.values_batch_update(
            {
                "valueInputOption": "RAW",
                "data": updates[i : i + chunk_size],
            }
        )

    print(f"  Wrote {len(fills)} platform values to sheet.")


def export_tracker_xlsx(credentials, out_path: str) -> None:
    """Export Sara's tracker from Google Drive as XLSX."""
    try:
        import requests
    except ImportError:
        print("Error: requests is not installed.")
        sys.exit(1)

    from google.auth.transport.requests import Request as GoogleRequest

    credentials.refresh(GoogleRequest())
    token = credentials.token

    url = DRIVE_EXPORT_URL.format(sheet_id=TRACKER_SHEET_ID)
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120)

    if resp.status_code != 200:
        print(f"Error: Drive export HTTP {resp.status_code} — {resp.text[:200]}")
        sys.exit(1)

    Path(out_path).write_bytes(resp.content)
    size_kb = len(resp.content) // 1024
    print(f"  Exported {out_path} ({size_kb} KB)")


def write_report(fills: list, total_rows: int, tarrow_total: int,
                 already_filled: int) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

    by_platform: dict = defaultdict(int)
    filled_list = []
    for _row, label in fills:
        by_platform[label] += 1
        filled_list.append({"row": _row, "platform": label})

    report = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "tracker_data_rows": total_rows,
        "tarrow_unique_urls": tarrow_total,
        "rows_already_filled": already_filled,
        "rows_newly_filled": len(fills),
        "by_platform": dict(by_platform),
        "fills": filled_list,
    }
    REPORT_PATH.write_text(json.dumps(report, indent=2, ensure_ascii=False))
    print(f"  Report: {REPORT_PATH}")


def main() -> int:
    try:
        import gspread
    except ImportError:
        print("Error: gspread is not installed. Run: pip install gspread")
        sys.exit(1)

    if not Path(TARROW_XLSX).exists():
        print(f"Error: {TARROW_XLSX} not found. Run download_tarrow.py first.")
        return 1

    credentials = _load_credentials()

    # ── Load Tarrow platform map ───────────────────────────────────────────────
    print("Loading Tarrow platform data…")
    tarrow_map = build_tarrow_platform_map(TARROW_XLSX)

    if not tarrow_map:
        print("Warning: Tarrow platform map is empty — nothing to backfill.")
        write_report([], 0, 0, 0)
        return 0

    # ── Open Sara's tracker ───────────────────────────────────────────────────
    print("Opening Sara's tracker…")
    gc = gspread.authorize(credentials)
    sh = gc.open_by_key(TRACKER_SHEET_ID)
    ws = sh.worksheet("Data")

    all_values = ws.get_all_values()
    if not all_values:
        print("Error: tracker sheet is empty.")
        return 1

    headers = all_values[0]
    try:
        url_idx = headers.index("Published URL/Link")
        plat_idx = headers.index("Syndication platform")
    except ValueError as exc:
        print(f"Error: expected column not found in tracker: {exc}")
        return 1

    total_data_rows = len(all_values) - 1
    plat_col_1based = plat_idx + 1  # gspread uses 1-based col numbers

    # Count rows that already have a platform value (for the report)
    already_filled = sum(
        1 for row in all_values[1:]
        if len(row) > plat_idx and row[plat_idx].strip()
    )
    print(f"  Tracker: {total_data_rows} data rows, {already_filled} already have platform")

    # ── Compute fills ─────────────────────────────────────────────────────────
    print("Computing fills…")
    fills = compute_fills(all_values, url_idx, plat_idx, tarrow_map)

    by_plat: dict = defaultdict(int)
    for _, label in fills:
        by_plat[label] += 1
    print(f"  Found {len(fills)} rows to fill: {dict(by_plat)}")

    # ── Write to sheet ────────────────────────────────────────────────────────
    if fills:
        print("Writing to Google Sheet…")
        apply_fills(ws, fills, plat_col_1based)
    else:
        print("No new platform data to backfill.")

    # ── Export tracker as XLSX ────────────────────────────────────────────────
    print(f"Exporting tracker as {TRACKER_XLSX_OUT}…")
    export_tracker_xlsx(credentials, TRACKER_XLSX_OUT)

    # ── Write report ──────────────────────────────────────────────────────────
    write_report(fills, total_data_rows, len(tarrow_map), already_filled)

    print(
        f"\n✓ Backfill complete: {len(fills)} rows filled, "
        f"{already_filled} already had platform, "
        f"{total_data_rows - already_filled - len(fills)} rows unmatched."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
